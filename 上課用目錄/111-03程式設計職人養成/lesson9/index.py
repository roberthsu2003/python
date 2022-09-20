
import tkinter as tk
import requests
from tkinter import messagebox
from tkinter import filedialog
import openpyxl

class Window(tk.Tk):
    def __init__(self,codes):
        super().__init__()
        self.codes = codes
        self.title("各縣市4天天氣預測")
        print(self.codes)
        tk.Label(self, text="各縣市4天天氣預測", font=("arial",20)).pack(padx=100, pady=20)

        buttonFrame = tk.Frame(self)        
        for index,cities in enumerate(self.codes.items()):            
            cname,ename= cities
            btn = tk.Button(buttonFrame,text=f"{cname}\n{ename}",padx=20,pady=5,width=5)
            btn.grid(column=index % 7,row=index // 7)
            btn.bind('<Button>',self.btnClick)
        buttonFrame.pack()

        
        self.chage_content_frame = tk.Frame(self) #會改變內容的frame       
        #這裏的內容必需使用者按下按鈕後才面產生
        self.chage_content_frame.pack()



    def btnClick(self,event):
        btn=event.widget
        btn_text = btn["text"]
        nameList = btn_text.split()
        cname = nameList[0]
        ename = nameList[1]
        if hasattr(self, 'displayFrame'):
            self.displayFrame.destroy()
        self.displayFrame = DisplayFrame(self.chage_content_frame,cname=cname,ename=ename)
        self.displayFrame.pack(padx=10,pady=10)
        

class DisplayFrame(tk.LabelFrame):
    def __init__(self,master,cname,ename):
        super().__init__(master)
        self.cname = cname
        self.ename = ename
        #建立title Label
        self.configure(text=f"{self.cname}-{self.ename}")
        try:
            self.forecast = DisplayFrame.get_forecast(ename=self.ename)
        except:
            messagebox.showerror('取得資料錯誤', '取得資料錯誤,請稍後再試')
            return
        #將資料切割為3等分,left_data,center_data,right_data
        tk.Button(self,text=f"{self.cname}天氣預報儲存為Excel",command=lambda:DisplayFrame.save_to_excel(self.forecast)).pack(anchor=tk.W,padx=10,pady=10)
        total_rows = len(self.forecast)
        columns = 3
        rows = total_rows // 3 + 1
        left_data = self.forecast[:rows]
        center_data = self.forecast[rows:rows*2]
        right_data = self.forecast[rows*2:]
        print(len(left_data))
        print(len(center_data))
        print(len(right_data))
        #建立3欄的顯示空間
        #左邊顯示空間建立      
        leftFrame = tk.Frame(self,width=150,height=150,borderwidth=1,relief=tk.GROOVE)
        tk.Label(leftFrame,text="日期-時間",background='#cccccc').grid(row=0,column=0,sticky=tk.W,padx=10,pady=3)
        tk.Label(leftFrame,text="溫度",background='#cccccc').grid(row=0,column=1,sticky=tk.W,padx=10,pady=3)
        tk.Label(leftFrame,text="狀態",background='#cccccc').grid(row=0,column=2,sticky=tk.W,padx=10,pady=3)
        tk.Label(leftFrame,text="濕度",background='#cccccc').grid(row=0,column=3,sticky=tk.W,padx=10,pady=3)
        for row_index,item in enumerate(left_data):
            for column_index,value in enumerate(item):
                tk.Label(leftFrame,text=value).grid(row=row_index+1,column=column_index,sticky=tk.W,padx=10,pady=3)
        leftFrame.pack(side=tk.LEFT,padx=10)

        #中間顯示空間建立
        centerFrame = tk.Frame(self,width=150,height=150,borderwidth=1,relief=tk.GROOVE)
        tk.Label(centerFrame,text="日期-時間",background='#cccccc').grid(row=0,column=0,sticky=tk.W,padx=10,pady=3)
        tk.Label(centerFrame,text="溫度",background='#cccccc').grid(row=0,column=1,sticky=tk.W,padx=10,pady=3)
        tk.Label(centerFrame,text="狀態",background='#cccccc').grid(row=0,column=2,sticky=tk.W,padx=10,pady=3)
        tk.Label(centerFrame,text="濕度",background='#cccccc').grid(row=0,column=3,sticky=tk.W,padx=10,pady=3)
        for row_index,item in enumerate(center_data):
            for column_index,value in enumerate(item):
                tk.Label(centerFrame,text=value).grid(row=row_index+1,column=column_index,sticky=tk.W,padx=10,pady=3)

        centerFrame.pack(side=tk.LEFT,padx=10)

        #右邊顯示空間建立
        rightFrame = tk.Frame(self,width=150,height=150,borderwidth=1,relief=tk.GROOVE)
        tk.Label(rightFrame,text="日期-時間",background='#cccccc').grid(row=0,column=0,sticky=tk.W,padx=10,pady=3)
        tk.Label(rightFrame,text="溫度",background='#cccccc').grid(row=0,column=1,sticky=tk.W,padx=10,pady=3)
        tk.Label(rightFrame,text="狀態",background='#cccccc').grid(row=0,column=2,sticky=tk.W,padx=10,pady=3)
        tk.Label(rightFrame,text="濕度",background='#cccccc').grid(row=0,column=3,sticky=tk.W,padx=10,pady=3)
        for row_index,item in enumerate(right_data):
            for column_index,value in enumerate(item):
                tk.Label(rightFrame,text=value).grid(row=row_index+1,column=column_index,sticky=tk.W,padx=10,pady=3)
        rightFrame.pack(side=tk.LEFT,anchor=tk.N,padx=10)

    @staticmethod
    def get_forecast(ename):        
        url = "https://api.openweathermap.org/data/2.5/forecast?q="+ename+",tw&APPID=29c4f184354b9889e87f7b494ac86aed&lang=zh_tw&units=metric"
        response = requests.request("GET",url)
        if response.ok == True:
            print("下載成功")    
            all_data = response.json()
        list_data = all_data['list']

        county_forcase = []
        for item in list_data:
            county_forcase.append([item['dt_txt'],item['main']['temp'],item['weather'][0]['description'],item['main']['humidity']])
        
        return county_forcase

    @staticmethod
    def save_to_excel(data):        
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "天氣預測表"
        sheet.cell(row=1, column=1, value="日期-時間")
        sheet.cell(row=1, column=2, value="溫度")
        sheet.cell(row=1, column=3, value="狀態")
        sheet.cell(row=1, column=4, value="濕度")
        for row_index,row_data in enumerate(data):
            for column_index,value in enumerate(row_data):
                sheet.cell(row=row_index+2,column=column_index+1,value=value)
        filePath = filedialog.askdirectory()
        wb.save(filePath+'/氣象預測.xlsx')
        messagebox.showinfo("ok",message="存檔成功")

        



def main():
    tw_county_names = {"台北":"Taipei",
                   "台中":"Taichung",
                   "基隆":"Keelung",
                   "台南":"Tainan",
                   "高雄":"Kaohsiung",
                   "新北":"New Taipei",
                   "宜蘭":"Yilan",
                   "桃園":"Taoyuan",
                   "嘉義":"Chiayi",
                   "新竹":"Hsinchu",
                   "苗栗":"Miaoli",
                   "南投":"Nantou",
                   "彰化":"Changhua",
                   "雲林":"Yunlin",
                   "屏東":"Pingtung",
                   "花蓮":"Hualien",
                   "台東":"Taitung",
                   "金門":"Kinmen",
                   "澎湖":"Penghu",
                   "連江":"Lienchiang"
                   }
    window = Window(codes=tw_county_names)
    window.mainloop()

if __name__ == "__main__":
    main()