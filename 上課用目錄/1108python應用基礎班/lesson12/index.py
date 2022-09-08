#!/usr/bin/python3
'''
#這個應用程式是做什麼的
#要注意什麼
'''
import datasource
import pandas as pd
from openpyxl import load_workbook

def doSomething(inputString):
    if inputString == '':
        return 0
    else:
        return int(inputString)

def main():
    data_list = datasource.get_pm25()
    dataFrame = pd.DataFrame(data=data_list,columns=["sitename","county","aqi","status","pm2.5","publishtime"])
    dataFrame["aqi"] = dataFrame["aqi"].map(doSomething)
    new_series = dataFrame["aqi"]
    zero_index = new_series[new_series==0].index
    dataFrame.drop(index=zero_index,inplace=True)
    delete_index = dataFrame[dataFrame["pm2.5"] == ""].index
    dataFrame.drop(index=delete_index,inplace=True)
    dataFrame['pm2.5'] = dataFrame['pm2.5'].astype(int)
    #取出aqi最大值
    aqiMax = dataFrame['aqi'].max()
    aqiMax_dataFrame = dataFrame.query("aqi>=@aqiMax")
    #取出aqi最小值
    aqiMin = dataFrame['aqi'].min()    
    aqiMin_dataFRame = dataFrame[dataFrame['aqi']  == aqiMin]

    dataFrame
    with pd.ExcelWriter('pm25.xlsx') as writer:
        dataFrame.to_excel(writer,sheet_name="all")
    
    wb = load_workbook('pm25.xlsx')
    sheet = wb['all']
    cell = sheet.cell(row=2,column=sheet.max_column+2)
    cell.value = "aqi最大值"

    for row in sheet.iter_rows(min_row=3, min_col=9,max_col=14, max_row=3):    
        for cellIndex,cell in enumerate(row):
            cell.value = aqiMax_dataFrame.columns[cellIndex]

    for row in sheet.iter_rows(min_row=4, min_col=9,max_col=14, max_row=4):    
        for cellIndex,cell in enumerate(row):
            cell.value = aqiMax_dataFrame.iloc[0,cellIndex]
    wb.save('pm25.xlsx')

    

if __name__ == "__main__":
    main()